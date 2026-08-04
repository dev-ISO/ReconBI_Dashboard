import os
import datetime
from openpyxl.utils import column_index_from_string, get_column_letter
from typing import List, Dict, Optional


def get_cell_value(sheet, row: int, col_letter: str):
    """
    Helper to read a cell value from the sheet.
    Returns None if the cell does not exist or is empty.
    """
    cell = sheet[f"{col_letter}{row}"]
    val = cell.value
    if val is None:
        return None
    
    # Force numeric or trimmed string
    if isinstance(val, (int, float)):
        return val
    return str(val).strip()



def convert_excel_date(val):
    """
    Convert an Excel serial date (numeric) to 'YYYY-MM-DD'.
    If not numeric, return as-is (string or None).
    """
    if not isinstance(val, (int, float)):
        return val
    # Excel "serial" date offset (assuming 1900 system)
    # Node code: (val - 25569) * 86400 * 1000 --> JS Date
    # In Python, we can do:
    date_origin = datetime.datetime(1899, 12, 30)  # 1 in Excel = 1899-12-31, so offset by 1 day
    date_obj = date_origin + datetime.timedelta(days=val)
    return date_obj.strftime("%Y-%m-%d")



def get_sheet_bounds(sheet):
    """
    Return the max row (1-based) and max col (1-based) from the sheet.
    Openpyxl provides min_row, max_row, min_column, max_column.
    """
    return {
        "maxRow": sheet.max_row,
        "maxCol": sheet.max_column,
    }



def col_letter_to_index(letter: str) -> int:
    """
    Convert column letter (e.g. 'A', 'CW') to numeric index (1-based).
    """
    return column_index_from_string(letter)



def col_index_to_letter(idx: int) -> str:
    """
    Convert 1-based numeric index to column letter.
    """
    return get_column_letter(idx)



def is_entire_column_blank(
    sheet, 
    col_letter: str, 
    row_start: int, 
    row_end: int
) -> bool:
    """
    Check if an entire column is blank for the rows [rowStart..rowEnd].
    Treat 0 as blank as well. If you want 0 to be considered data, remove
    the `val != 0` check.
    """
    for row_num in range(row_start, row_end + 1):
        val = get_cell_value(sheet, row_num, col_letter)
        # If val is anything but None, "" or 0, we consider it "data"
        if val not in (None, "", 0):
            return False
    return True


# --------------------------------------------------
# 2) TABLE-PARSING LOGIC
# --------------------------------------------------


def process_single_table(
    sheet, 
    header_row: int, 
    start_col_letter: str
) -> List[Dict]:
    """
    Read a single table starting at `header_row` (the row with "Office" in col A).
      - The row after that is the data start.
      - Keep reading rows until we find "Total" in col C (include that row).
      - For columns, start at `start_col_letter` and go right until we find 
        the first column that is entirely blank for all data rows.
    """
    results = []

    # Data begins on the next row
    data_start_row = header_row + 1

    # 1) Find the row block for this table
    current_row = data_start_row
    table_end_row = data_start_row - 1

    while True:
        bounds = get_sheet_bounds(sheet)
        max_row = bounds["maxRow"]
        if current_row > max_row:
            break

        office_val = get_cell_value(sheet, current_row, "A")
        resource_val = get_cell_value(sheet, current_row, "C")

        # If col A or col C is empty => end of table
        if not office_val or not resource_val:
            break

        table_end_row = current_row

        # If "Total", include that row and stop
        if isinstance(resource_val, str) and resource_val.lower() == "total":
            break

        current_row += 1

    # If no data
    if table_end_row < data_start_row:
        return results

    # 2) Figure out how far to the right to read columns
    bounds = get_sheet_bounds(sheet)
    max_col = bounds["maxCol"]

    col_index = col_letter_to_index(start_col_letter)
    valid_columns = []

    while col_index <= max_col:
        col_letter = col_index_to_letter(col_index)
        if is_entire_column_blank(sheet, col_letter, data_start_row, table_end_row):
            break
        valid_columns.append(col_letter)
        col_index += 1

    # 3) Build array of headers
    headers = []
    for col_letter in valid_columns:
        raw = get_cell_value(sheet, header_row, col_letter)
        headers.append({
            "col": col_letter,
            "date": convert_excel_date(raw)  # might be numeric date, might be string
        })

    # 4) Collect the row data
    for row_num in range(data_start_row, table_end_row + 1):
        office_val = get_cell_value(sheet, row_num, "A")
        resource_val = get_cell_value(sheet, row_num, "C")

        if not office_val or not resource_val:
            break

        row_data = {
            "office": str(office_val).strip(),
            "resource": str(resource_val).strip(),
            "data": []
        }

        # For each valid column, read the cell
        for hdr in headers:
            col = hdr["col"]
            date_str = hdr["date"]
            val = get_cell_value(sheet, row_num, col)
            # If cell is None, store as None
            row_data["data"].append({
                "date": date_str,
                "value": val if val is not None else None
            })

        results.append(row_data)

        # Stop after "Total"
        if isinstance(resource_val, str) and resource_val.lower() == "total":
            break

    return results



def process_all_tables(sheet, start_col_letter="CW") -> List[Dict]:
    """
    Main function:
      - Scan down column A looking for the word "Office".
      - Each time we find it, call `process_single_table` with that row as header row.
    """
    results = []
    bounds = get_sheet_bounds(sheet)
    max_row = bounds["maxRow"]

    current_row = 1
    while current_row <= max_row:
        val_a = get_cell_value(sheet, current_row, "A")

        if isinstance(val_a, str) and "office" in val_a.lower():
            table_data = process_single_table(sheet, current_row, start_col_letter)
            results.extend(table_data)

            # Move current_row beyond that table
            if len(table_data) > 0:
                jump_row = current_row + 1
                while jump_row <= max_row:
                    maybe_office = get_cell_value(sheet, jump_row, "A")
                    maybe_resource = get_cell_value(sheet, jump_row, "C")
                    if not maybe_office or not maybe_resource:
                        jump_row += 1
                        break
                    if isinstance(maybe_office, str) and "office" in maybe_office.lower():
                        break
                    jump_row += 1
                current_row = jump_row
            else:
                current_row += 1
        else:
            current_row += 1

    return results